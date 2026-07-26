import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.estoque.models import LocalEstoque
from apps.financeiro.models import ParceiroFinanceiro
from apps.talhoes.models import Talhao

from .grain_access import cadpros_visiveis, exigir_acesso_cadpro
from .grain_models import (
    CadPro,
    ContratoProducao,
    Cultura,
    EmbarqueProducao,
    ImportacaoPlanilha,
    Motorista,
    MovimentacaoGraos,
    RecebimentoProducao,
    Safra,
    Veiculo,
)
from .grain_services import registrar_auditoria, registrar_movimentacao


MAX_IMPORT_SIZE = 10 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
MAX_PREVIEW_ROWS = 100


FIELD_ALIASES = {
    "data": {"data", "data recebimento", "data da retirada", "emissao"},
    "motorista": {"motorista", "nome motorista"},
    "placa": {"placa", "placas", "placa veiculo", "veiculo"},
    "cadpro": {"cadpro", "cad pro", "cad/pro", "inscricao produtor"},
    "talhao": {"talhao", "talhão", "lote origem"},
    "cultura": {"cultura", "produto", "grao", "grão"},
    "safra": {"safra", "cultura safra"},
    "peso_bruto_kg": {"peso bruto", "peso bruto kg", "bruto kg"},
    "tara_kg": {"tara", "tara kg"},
    "peso_liquido_kg": {"peso liquido", "peso líquido", "peso liquido kg", "liquido kg"},
    "umidade_percentual": {"umidade", "umidade %", "percentual umidade"},
    "impureza_percentual": {"impureza", "impurezas", "impureza %"},
    "defeitos_percentual": {"defeitos", "defeito", "defeitos %"},
    "local_armazenagem": {"local", "armazenagem", "silo", "armazem", "armazém"},
    "romaneio": {"romaneio", "numero romaneio", "n romaneio"},
    "comprador": {"comprador", "cliente", "destinatario", "destinatário"},
    "contrato": {"contrato", "numero contrato", "n contrato"},
    "nota_produtor": {"nota produtor", "nfe produtor", "nf produtor", "nº da nfe"},
    "nota_empresa": {"nota empresa", "nfe empresa", "nf empresa"},
    "destino": {"destino", "local destino"},
    "quantidade_kg": {"quantidade kg", "qtd kg", "kg", "qt em litros ou kg"},
    "preco_saca": {"preco", "preço", "preco saca", "valor saca"},
    "tipo_movimentacao": {"tipo", "tipo movimentacao", "movimento"},
    "local_origem": {"origem", "local origem"},
    "local_destino": {"destino estoque", "local destino estoque"},
    "motivo": {"motivo", "observacao", "observação"},
}


REQUIRED_FIELDS = {
    ImportacaoPlanilha.Tipo.RECEBIMENTOS: {
        "data",
        "cultura",
        "safra",
        "peso_bruto_kg",
        "peso_liquido_kg",
        "local_armazenagem",
    },
    ImportacaoPlanilha.Tipo.MOVIMENTACOES: {
        "tipo_movimentacao",
        "cultura",
        "safra",
        "quantidade_kg",
    },
    ImportacaoPlanilha.Tipo.EMBARQUES: {
        "data",
        "comprador",
        "cultura",
        "safra",
        "local_armazenagem",
        "romaneio",
        "quantidade_kg",
        "preco_saca",
    },
}


def normalize_text(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-zA-Z0-9%/]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def file_hash(uploaded_file):
    digest = hashlib.sha256()
    position = uploaded_file.tell()
    uploaded_file.seek(0)
    for chunk in iter(lambda: uploaded_file.read(1024 * 1024), b""):
        digest.update(chunk)
    uploaded_file.seek(position)
    return digest.hexdigest()


def validate_upload(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("Formato não suportado. Use CSV, XLSX ou XLSM.")
    if uploaded_file.size > MAX_IMPORT_SIZE:
        raise ValueError("A planilha excede o limite de 10 MB.")
    return extension


def _read_csv(uploaded_file):
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return list(csv.reader(io.StringIO(text), dialect=dialect))


def _read_workbook(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def read_rows(uploaded_file):
    extension = validate_upload(uploaded_file)
    rows = _read_csv(uploaded_file) if extension == ".csv" else _read_workbook(uploaded_file)
    while rows and not any(value not in (None, "") for value in rows[0]):
        rows.pop(0)
    if not rows:
        raise ValueError("A planilha está vazia.")
    headers = [str(value or "").strip() for value in rows[0]]
    data = []
    for index, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[column]: row[column] if column < len(row) else None for column in range(len(headers))}
        record["__linha__"] = index
        data.append(record)
    return headers, data


def detect_mapping(headers):
    normalized = {header: normalize_text(header) for header in headers}
    mapping = {}
    for field, aliases in FIELD_ALIASES.items():
        aliases_normalized = {normalize_text(alias) for alias in aliases}
        for header, value in normalized.items():
            if value in aliases_normalized:
                mapping[field] = header
                break
    return mapping


def _serialize_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _mapped_row(row, mapping):
    return {
        field: _serialize_value(row.get(column))
        for field, column in mapping.items()
        if column
    }


def _decimal(value, field, line, errors, required=False):
    if value in (None, ""):
        if required:
            errors.append({"linha": line, "campo": field, "mensagem": "Valor obrigatório."})
        return None
    try:
        normalized = str(value).strip().replace(".", "").replace(",", ".") if isinstance(value, str) and "," in value else str(value)
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        errors.append({"linha": line, "campo": field, "mensagem": "Número inválido."})
        return None


def _date(value, line, errors):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if not value:
        errors.append({"linha": line, "campo": "data", "mensagem": "Data obrigatória."})
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    errors.append({"linha": line, "campo": "data", "mensagem": "Data inválida."})
    return None


def preview_import(uploaded_file, import_type, manual_mapping=None):
    headers, rows = read_rows(uploaded_file)
    mapping = {**detect_mapping(headers), **(manual_mapping or {})}
    missing = sorted(REQUIRED_FIELDS[import_type] - set(mapping))
    errors = [
        {"linha": 1, "campo": field, "mensagem": "Coluna obrigatória não mapeada."}
        for field in missing
    ]
    preview = []
    for row in rows[:MAX_PREVIEW_ROWS]:
        line = row["__linha__"]
        item = _mapped_row(row, mapping)
        item["__linha__"] = line
        if import_type == ImportacaoPlanilha.Tipo.RECEBIMENTOS:
            _date(item.get("data"), line, errors)
            for field in ("peso_bruto_kg", "peso_liquido_kg"):
                _decimal(item.get(field), field, line, errors, required=True)
        elif import_type == ImportacaoPlanilha.Tipo.MOVIMENTACOES:
            _decimal(item.get("quantidade_kg"), "quantidade_kg", line, errors, required=True)
        else:
            _date(item.get("data"), line, errors)
            _decimal(item.get("quantidade_kg"), "quantidade_kg", line, errors, required=True)
            _decimal(item.get("preco_saca"), "preco_saca", line, errors, required=True)
        preview.append(item)
    return {
        "headers": headers,
        "mapping": mapping,
        "preview": preview,
        "errors": errors,
        "total_rows": len(rows),
    }


def _get_cadpro(import_record, row):
    if import_record.cadpro_id:
        return import_record.cadpro
    code = str(row.get("cadpro") or "").strip()
    if not code:
        raise ValueError("Informe o CAD/PRO na importação ou na planilha.")
    cadpro = cadpros_visiveis(import_record.criado_por, import_record.propriedade_id).filter(codigo__iexact=code).first()
    if not cadpro:
        raise ValueError(f"CAD/PRO não encontrado: {code}.")
    return cadpro


def _get_culture(value):
    text = str(value or "").strip()
    culture = Cultura.objects.filter(nome__iexact=text).first() or Cultura.objects.filter(codigo__iexact=normalize_text(text).replace(" ", "-")).first()
    if not culture:
        raise ValueError(f"Cultura não cadastrada: {text}.")
    return culture


def _get_harvest(value):
    text = str(value or "").strip()
    harvest = Safra.objects.filter(nome__iexact=text).first()
    if not harvest:
        raise ValueError(f"Safra não cadastrada: {text}.")
    return harvest


def _get_location(value, property_id):
    text = str(value or "").strip()
    location = LocalEstoque.objects.filter(nome__iexact=text, propriedade_id__in=[property_id, None]).first()
    if not location:
        raise ValueError(f"Local de armazenagem não cadastrado: {text}.")
    return location


def _get_field(value, property_id):
    text = str(value or "").strip()
    if not text:
        return None
    field = Talhao.objects.filter(nome__iexact=text, propriedade_id=property_id).first()
    if not field:
        raise ValueError(f"Talhão não cadastrado: {text}.")
    return field


def _get_driver(value):
    text = str(value or "").strip()
    if not text:
        return None
    return Motorista.objects.get_or_create(nome=text, defaults={"ativo": True})[0]


def _get_vehicle(value):
    text = str(value or "").strip()
    if not text:
        return None
    plate = text.replace(" ", "").replace("-", "").upper()[:10]
    return Veiculo.objects.get_or_create(placa=plate, defaults={"ativo": True})[0]


def _get_buyer(value):
    text = str(value or "").strip()
    buyer = ParceiroFinanceiro.objects.filter(nome__iexact=text).first()
    if not buyer:
        raise ValueError(f"Comprador não cadastrado: {text}.")
    return buyer


def _row_records(import_record):
    headers, rows = read_rows(import_record.arquivo)
    del headers
    return [_mapped_row(row, import_record.mapeamento) | {"__linha__": row["__linha__"]} for row in rows]


@transaction.atomic
def confirm_import(import_record, *, user):
    import_record = ImportacaoPlanilha.objects.select_for_update().select_related("propriedade", "cadpro").get(pk=import_record.pk)
    if import_record.criado_por_id != user.id and not user.is_superuser:
        raise PermissionError("A importação pertence a outro usuário.")
    if import_record.status != ImportacaoPlanilha.Status.VALIDADA:
        raise ValueError("Somente importações validadas podem ser confirmadas.")
    if import_record.cadpro:
        exigir_acesso_cadpro(user, import_record.cadpro)

    imported = 0
    issues = []
    for row in _row_records(import_record):
        line = row["__linha__"]
        try:
            cadpro = _get_cadpro(import_record, row)
            exigir_acesso_cadpro(user, cadpro)
            culture = _get_culture(row.get("cultura"))
            harvest = _get_harvest(row.get("safra"))
            if import_record.tipo == ImportacaoPlanilha.Tipo.RECEBIMENTOS:
                receipt = RecebimentoProducao.objects.create(
                    data=_date(row.get("data"), line, []),
                    propriedade=import_record.propriedade,
                    cadpro=cadpro,
                    talhao=_get_field(row.get("talhao"), import_record.propriedade_id),
                    cultura=culture,
                    safra=harvest,
                    local_armazenagem=_get_location(row.get("local_armazenagem"), import_record.propriedade_id),
                    motorista=_get_driver(row.get("motorista")),
                    veiculo=_get_vehicle(row.get("placa")),
                    placa_informada=str(row.get("placa") or ""),
                    romaneio=str(row.get("romaneio") or ""),
                    peso_bruto_kg=_decimal(row.get("peso_bruto_kg"), "peso_bruto_kg", line, [], True),
                    tara_kg=_decimal(row.get("tara_kg") or 0, "tara_kg", line, []),
                    peso_liquido_kg=_decimal(row.get("peso_liquido_kg"), "peso_liquido_kg", line, [], True),
                    umidade_percentual=_decimal(row.get("umidade_percentual") or 0, "umidade_percentual", line, []),
                    impureza_percentual=_decimal(row.get("impureza_percentual") or 0, "impureza_percentual", line, []),
                    defeitos_percentual=_decimal(row.get("defeitos_percentual") or 0, "defeitos_percentual", line, []),
                    observacoes="Importado de planilha; pendente de confirmação operacional.",
                    criado_por=user,
                )
                receipt.full_clean()
                receipt.save()
                registrar_auditoria(usuario=user, acao="recebimento_importado", objeto=receipt, metadados={"importacao": import_record.pk, "linha": line})
            elif import_record.tipo == ImportacaoPlanilha.Tipo.MOVIMENTACOES:
                type_map = {
                    "entrada": MovimentacaoGraos.Tipo.AJUSTE_ENTRADA,
                    "saida": MovimentacaoGraos.Tipo.AJUSTE_SAIDA,
                    "saída": MovimentacaoGraos.Tipo.AJUSTE_SAIDA,
                    "transferencia": MovimentacaoGraos.Tipo.TRANSFERENCIA,
                    "transferência": MovimentacaoGraos.Tipo.TRANSFERENCIA,
                }
                movement_type = type_map.get(str(row.get("tipo_movimentacao") or "").strip().lower())
                if not movement_type:
                    raise ValueError("Tipo de movimentação inválido.")
                origin = _get_location(row.get("local_origem"), import_record.propriedade_id) if row.get("local_origem") else None
                destination = _get_location(row.get("local_destino"), import_record.propriedade_id) if row.get("local_destino") else None
                registrar_movimentacao(
                    usuario=user,
                    tipo=movement_type,
                    propriedade=import_record.propriedade,
                    cadpro=cadpro,
                    talhao=_get_field(row.get("talhao"), import_record.propriedade_id),
                    cultura=culture,
                    safra=harvest,
                    quantidade_kg=_decimal(row.get("quantidade_kg"), "quantidade_kg", line, [], True),
                    local_origem=origin,
                    local_destino=destination,
                    referencia_tipo="importacao",
                    referencia_id=import_record.pk,
                    motivo=str(row.get("motivo") or "Importado de planilha"),
                )
            else:
                buyer = _get_buyer(row.get("comprador"))
                contract_number = str(row.get("contrato") or "").strip()
                contract = ContratoProducao.objects.filter(
                    comprador=buyer,
                    numero__iexact=contract_number,
                ).first() if contract_number else None
                shipment = EmbarqueProducao.objects.create(
                    data=_date(row.get("data"), line, []),
                    propriedade=import_record.propriedade,
                    cadpro=cadpro,
                    cultura=culture,
                    safra=harvest,
                    local_armazenagem=_get_location(row.get("local_armazenagem"), import_record.propriedade_id),
                    comprador=buyer,
                    contrato=contract,
                    motorista=_get_driver(row.get("motorista")),
                    veiculo=_get_vehicle(row.get("placa")),
                    placa_informada=str(row.get("placa") or ""),
                    destino=str(row.get("destino") or ""),
                    romaneio=str(row.get("romaneio") or "").strip(),
                    nota_produtor=str(row.get("nota_produtor") or ""),
                    nota_empresa=str(row.get("nota_empresa") or ""),
                    quantidade_kg=_decimal(row.get("quantidade_kg"), "quantidade_kg", line, [], True),
                    preco_saca=_decimal(row.get("preco_saca"), "preco_saca", line, [], True),
                    observacoes="Importado de planilha; pendente de confirmação operacional.",
                    criado_por=user,
                )
                shipment.full_clean()
                shipment.save()
                registrar_auditoria(usuario=user, acao="embarque_importado", objeto=shipment, metadados={"importacao": import_record.pk, "linha": line})
            imported += 1
        except Exception as exc:
            issues.append({"linha": line, "mensagem": str(exc)})

    if issues:
        raise ValueError({"detail": "A importação não foi aplicada porque existem inconsistências.", "inconsistencias": issues})
    import_record.linhas_importadas = imported
    import_record.status = ImportacaoPlanilha.Status.IMPORTADA
    import_record.confirmado_em = timezone.now()
    import_record.save(update_fields=("linhas_importadas", "status", "confirmado_em"))
    registrar_auditoria(usuario=user, acao="planilha_importada", objeto=import_record, metadados={"linhas": imported})
    return import_record
