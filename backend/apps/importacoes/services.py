import hashlib
import json
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from django.db import IntegrityError, transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.graos.models import LoteGraos
from apps.propriedades.models import Propriedade

from .models import LinhaImportacao, LoteImportacao


MAX_ARQUIVO_BYTES = 10 * 1024 * 1024
MAX_CONTEUDO_DESCOMPACTADO_BYTES = 100 * 1024 * 1024
MAX_ENTRADAS_ZIP = 5000
PLANILHAS_PRODUCAO = {str(numero) for numero in range(1, 46)}
CABECALHOS_OBRIGATORIOS = {
    LinhaImportacao.Tipo.PRODUCAO: {
        "B6": "DATA",
        "D6": "PESO (KG)",
        "K6": "PESO LIQUIDO (KG)",
    },
    LinhaImportacao.Tipo.SAIDA: {
        "B6": "DATA",
        "C6": "DESTINO",
        "G6": "CADPRO",
        "L6": "PESO LIQUIDO (KG)",
    },
    LinhaImportacao.Tipo.TERCEIROS: {
        "B6": "DATA",
        "C6": "PRODUTOR",
        "E6": "PESO (KG)",
        "K6": "PESO LIQUIDO (KG)",
    },
}


class PlanilhaImportacaoError(ValueError):
    pass


class ArquivoImportacaoDuplicadoError(PlanilhaImportacaoError):
    def __init__(self, lote):
        self.lote = lote
        super().__init__(
            f"Este arquivo já foi importado no lote {lote.id} "
            f"({lote.arquivo_sha256})."
        )


def normalizar_texto(valor):
    if valor in (None, ""):
        return ""
    return " ".join(str(valor).strip().split())


def chave_texto(valor):
    texto = unicodedata.normalize("NFKD", normalizar_texto(valor))
    return "".join(
        caractere for caractere in texto if not unicodedata.combining(caractere)
    ).upper()


def valor_json(valor):
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    if isinstance(valor, Decimal):
        return str(valor)
    if isinstance(valor, float):
        return str(Decimal(str(valor)))
    if valor is None or isinstance(valor, (str, int, bool)):
        return valor
    return str(valor)


def decimal_ou_none(valor):
    if valor in (None, "") or isinstance(valor, bool):
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        if "," in valor:
            valor = valor.replace(".", "").replace(",", ".")
    try:
        return Decimal(str(valor))
    except (InvalidOperation, TypeError, ValueError):
        return None


def data_ou_none(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = normalizar_texto(valor)
    if not texto:
        return None
    for formato in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def normalizar_safra(valor):
    texto = normalizar_texto(valor)
    advertencias = []
    if not texto:
        return "", ["Safra ausente na planilha."]
    texto = texto.replace(" ", "")
    match = re.fullmatch(r"(\d{2})/(\d{2})", texto)
    if match:
        inicio = 2000 + int(match.group(1))
        fim = 2000 + int(match.group(2))
        if fim != inicio + 1:
            advertencias.append("Safra possui intervalo não consecutivo.")
        return f"{inicio:04d}/{fim:04d}", advertencias
    match = re.fullmatch(r"(\d{4})/(\d{4})", texto)
    if match:
        inicio, fim = int(match.group(1)), int(match.group(2))
        if fim != inicio + 1:
            advertencias.append("Safra possui intervalo não consecutivo.")
        return f"{inicio:04d}/{fim:04d}", advertencias
    if re.fullmatch(r"\d{4}(?:\.0+)?", texto):
        inicio = int(float(texto))
        advertencias.append(
            "Safra informada com um único ano; interpretada como ano agrícola."
        )
        return f"{inicio:04d}/{inicio + 1:04d}", advertencias
    return texto, ["Formato de safra atípico; valor preservado."]


def normalizar_cadpro(valor):
    texto = normalizar_texto(valor)
    correspondencia = re.search(r"(?<!\d)(\d{7,})(?!\d)", texto)
    return correspondencia.group(1) if correspondencia else ""


def formula_ou_none(sheet, coordenada):
    valor = sheet[coordenada].value
    if isinstance(valor, str) and valor.startswith("="):
        return valor
    return None


def _validar_cabecalhos(sheet, tipo):
    encontrados = {}
    ausentes = []
    for coordenada, esperado in CABECALHOS_OBRIGATORIOS[tipo].items():
        valor = normalizar_texto(sheet[coordenada].value)
        encontrados[coordenada] = valor
        if chave_texto(valor) != chave_texto(esperado):
            ausentes.append(f"{coordenada}={esperado}")
    if ausentes:
        raise PlanilhaImportacaoError(
            f"A aba {sheet.title} nao possui os cabecalhos obrigatorios: "
            f"{', '.join(ausentes)}."
        )
    return encontrados


def _validar_percentual(nome, valor, erros):
    if valor is not None and not Decimal("0") <= valor <= Decimal("100"):
        erros.append(f"{nome} deve estar entre 0 e 100.")


class AssociadorGraos:
    def __init__(self):
        self.propriedades = {}
        for propriedade in Propriedade.objects.all():
            self.propriedades.setdefault(chave_texto(propriedade.nome), []).append(
                propriedade
            )
        self.lotes = {}
        queryset = LoteGraos.objects.select_related("armazem__propriedade")
        for lote in queryset:
            chave = (
                lote.armazem.propriedade_id,
                chave_texto(lote.cultura),
                lote.safra,
            )
            self.lotes.setdefault(chave, []).append(lote)

    def associar(self, *, propriedade_nome, cultura, safra):
        advertencias = []
        propriedade = None
        lote_graos = None
        chave_completa = chave_texto(propriedade_nome)
        chaves = [chave_completa]
        sem_cadpro = re.sub(r"\s*-\s*\d{7,}$", "", chave_completa).strip()
        if sem_cadpro and sem_cadpro != chave_completa:
            chaves.append(sem_cadpro)

        candidatas = []
        for chave in chaves:
            candidatas.extend(self.propriedades.get(chave, []))
        candidatas = list({item.id: item for item in candidatas}.values())
        if len(candidatas) == 1:
            propriedade = candidatas[0]
        elif len(candidatas) > 1:
            advertencias.append(
                "Associação de propriedade ambígua; revisão manual necessária."
            )
        else:
            advertencias.append(
                "Propriedade não associada ao cadastro; revisão manual necessária."
            )

        if propriedade:
            lotes = self.lotes.get(
                (propriedade.id, chave_texto(cultura), safra),
                [],
            )
            if len(lotes) == 1:
                lote_graos = lotes[0]
            elif len(lotes) > 1:
                advertencias.append(
                    "Mais de um lote de grãos corresponde ao contexto informado."
                )
            else:
                advertencias.append(
                    "Nenhum lote de grãos corresponde à propriedade, cultura e safra."
                )
        return propriedade, lote_graos, advertencias


def _linha_base(
    *,
    planilha,
    linha_origem,
    tipo,
    originais,
    normalizados,
    erros,
    advertencias,
    associador,
):
    propriedade, lote_graos, avisos_associacao = associador.associar(
        propriedade_nome=normalizados.get("propriedade_nome", ""),
        cultura=normalizados.get("cultura", ""),
        safra=normalizados.get("safra", ""),
    )
    advertencias.extend(avisos_associacao)
    status = LinhaImportacao.Status.ERRO if erros else (
        LinhaImportacao.Status.ADVERTENCIA
        if advertencias
        else LinhaImportacao.Status.VALIDA
    )
    associacao = LinhaImportacao.Associacao.NAO_ASSOCIADA
    if lote_graos:
        associacao = LinhaImportacao.Associacao.LOTE_GRAOS
    elif propriedade:
        associacao = LinhaImportacao.Associacao.PROPRIEDADE
    originais_json = {
        chave: valor_json(valor) for chave, valor in originais.items()
    }
    normalizados_json = {
        chave: valor_json(valor) for chave, valor in normalizados.items()
    }
    conteudo_hash = json.dumps(
        {
            "tipo": tipo,
            "dados": normalizados_json,
        },
        sort_keys=True,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return {
        "planilha": planilha,
        "linha_origem": linha_origem,
        "tipo": tipo,
        "status": status,
        "hash_linha": hashlib.sha256(conteudo_hash).hexdigest(),
        "dados_originais": originais_json,
        "dados_normalizados": normalizados_json,
        "erros": erros,
        "advertencias": advertencias,
        "associacao": associacao,
        "propriedade": propriedade,
        "lote_graos": lote_graos,
    }


def _parse_producao(sheet, formula_sheet, associador):
    propriedade_nome = normalizar_texto(sheet["C2"].value)
    cultura = normalizar_texto(sheet["F3"].value).title()
    safra, avisos_safra = normalizar_safra(sheet["C4"].value)
    resultados = []
    for numero, celulas in enumerate(
        sheet.iter_rows(min_row=7, max_col=11),
        start=7,
    ):
        data_valor = celulas[1].value
        peso_valor = celulas[3].value
        if data_valor in (None, "") and peso_valor in (None, "", 0):
            continue
        originais = {
            "propriedade": sheet["C2"].value,
            "area": sheet["C3"].value,
            "cultura": sheet["F3"].value,
            "safra": sheet["C4"].value,
            "data": data_valor,
            "identificacao_transporte": celulas[2].value,
            "peso_bruto_kg": peso_valor,
            "umidade_percentual": celulas[4].value,
            "impureza_percentual": celulas[5].value,
            "defeitos_percentual": celulas[6].value,
            "ph": celulas[7].value,
            "semente": celulas[8].value,
            "local_armazenagem": celulas[9].value,
            "peso_liquido_kg": celulas[10].value,
            "peso_liquido_formula": formula_ou_none(
                formula_sheet,
                f"K{numero}",
            ),
        }
        data = data_ou_none(data_valor)
        peso_bruto = decimal_ou_none(peso_valor)
        peso_liquido = decimal_ou_none(celulas[10].value)
        umidade = decimal_ou_none(celulas[4].value)
        impureza = decimal_ou_none(celulas[5].value)
        defeitos = decimal_ou_none(celulas[6].value)
        erros = []
        advertencias = list(avisos_safra)
        if not propriedade_nome or chave_texto(propriedade_nome) == "EM BRANCO":
            erros.append("Propriedade ausente ou identificada como EM BRANCO.")
        if not cultura:
            erros.append("Cultura ausente.")
        if not data:
            erros.append("Data inválida ou ausente.")
        if peso_bruto is None or peso_bruto <= 0:
            erros.append("Peso bruto deve ser maior que zero.")
        if peso_liquido is None or peso_liquido <= 0:
            erros.append("Peso líquido calculado está ausente ou inválido.")
        elif peso_bruto is not None and peso_liquido > peso_bruto:
            erros.append("Peso líquido não pode superar o peso bruto.")
        _validar_percentual("Umidade", umidade, erros)
        _validar_percentual("Impureza", impureza, erros)
        _validar_percentual("Defeitos", defeitos, erros)
        normalizados = {
            "propriedade_nome": propriedade_nome,
            "area": decimal_ou_none(sheet["C3"].value),
            "cultura": cultura,
            "safra": safra,
            "cadpro_numero": normalizar_cadpro(propriedade_nome),
            "classificacao_codigo": "PADRAO",
            "data": data,
            "identificacao_transporte": normalizar_texto(
                celulas[2].value
            ).upper(),
            "peso_bruto_kg": peso_bruto,
            "peso_liquido_kg": peso_liquido,
            "umidade_percentual": umidade,
            "impureza_percentual": impureza,
            "defeitos_percentual": defeitos,
            "ph": decimal_ou_none(celulas[7].value),
            "destinado_semente": bool(
                normalizar_texto(celulas[8].value)
            ),
            "local_armazenagem": normalizar_texto(
                celulas[9].value
            ),
        }
        resultados.append(
            _linha_base(
                planilha=sheet.title,
                linha_origem=numero,
                tipo=LinhaImportacao.Tipo.PRODUCAO,
                originais=originais,
                normalizados=normalizados,
                erros=erros,
                advertencias=advertencias,
                associador=associador,
            )
        )
    return resultados


def _parse_saida(sheet, formula_sheet, associador):
    cultura = normalizar_texto(sheet["C3"].value).title()
    safra, avisos_safra = normalizar_safra(sheet["C4"].value)
    resultados = []
    for numero, celulas in enumerate(
        sheet.iter_rows(min_row=7, max_col=12),
        start=7,
    ):
        data_valor = celulas[1].value
        peso_valor = celulas[11].value
        if data_valor in (None, "") and peso_valor in (None, "", 0):
            continue
        propriedade_nome = normalizar_texto(celulas[6].value)
        originais = {
            "cultura": sheet["C3"].value,
            "safra": sheet["C4"].value,
            "data": data_valor,
            "destino": celulas[2].value,
            "placa": celulas[5].value,
            "cadpro": celulas[6].value,
            "produtor": celulas[7].value,
            "contrato": celulas[8].value,
            "nota_produtor": celulas[9].value,
            "nota_empresa": celulas[10].value,
            "peso_liquido_kg": peso_valor,
            "peso_liquido_formula": formula_ou_none(
                formula_sheet,
                f"L{numero}",
            ),
        }
        data = data_ou_none(data_valor)
        peso_liquido = decimal_ou_none(peso_valor)
        erros = []
        advertencias = list(avisos_safra)
        if not data:
            erros.append("Data inválida ou ausente.")
        if peso_liquido is None or peso_liquido <= 0:
            erros.append("Peso líquido deve ser maior que zero.")
        if not normalizar_texto(celulas[2].value):
            erros.append("Destino ausente.")
        if not propriedade_nome:
            erros.append("CADPRO ou propriedade de origem ausente.")
        if not normalizar_texto(celulas[8].value):
            advertencias.append("Número de contrato ausente.")
        normalizados = {
            "propriedade_nome": propriedade_nome,
            "cultura": cultura,
            "safra": safra,
            "cadpro_numero": normalizar_cadpro(celulas[6].value),
            "classificacao_codigo": "PADRAO",
            "data": data,
            "destino": normalizar_texto(celulas[2].value),
            "placa": normalizar_texto(celulas[5].value).upper(),
            "produtor": normalizar_texto(celulas[7].value),
            "contrato": normalizar_texto(celulas[8].value),
            "nota_produtor": normalizar_texto(celulas[9].value),
            "nota_empresa": normalizar_texto(celulas[10].value),
            "peso_liquido_kg": peso_liquido,
        }
        resultados.append(
            _linha_base(
                planilha=sheet.title,
                linha_origem=numero,
                tipo=LinhaImportacao.Tipo.SAIDA,
                originais=originais,
                normalizados=normalizados,
                erros=erros,
                advertencias=advertencias,
                associador=associador,
            )
        )
    return resultados


def _parse_terceiros(sheet, formula_sheet, associador):
    cultura = normalizar_texto(sheet["C3"].value).title()
    safra, avisos_safra = normalizar_safra(sheet["C4"].value)
    resultados = []
    for numero, celulas in enumerate(
        sheet.iter_rows(min_row=7, max_col=11),
        start=7,
    ):
        data_valor = celulas[1].value
        peso_valor = celulas[4].value
        if data_valor in (None, "") and peso_valor in (None, "", 0):
            continue
        produtor = normalizar_texto(celulas[2].value)
        originais = {
            "cultura": sheet["C3"].value,
            "safra": sheet["C4"].value,
            "data": data_valor,
            "produtor": celulas[2].value,
            "peso_bruto_kg": peso_valor,
            "umidade_percentual": celulas[5].value,
            "impureza_percentual": celulas[6].value,
            "defeitos_percentual": celulas[7].value,
            "ph": celulas[8].value,
            "semente": celulas[9].value,
            "peso_liquido_kg": celulas[10].value,
            "peso_liquido_formula": formula_ou_none(
                formula_sheet,
                f"K{numero}",
            ),
        }
        data = data_ou_none(data_valor)
        peso_bruto = decimal_ou_none(peso_valor)
        peso_liquido = decimal_ou_none(celulas[10].value)
        umidade = decimal_ou_none(celulas[5].value)
        impureza = decimal_ou_none(celulas[6].value)
        defeitos = decimal_ou_none(celulas[7].value)
        erros = []
        advertencias = list(avisos_safra)
        if not produtor:
            erros.append("Produtor terceiro ausente.")
        if not data:
            erros.append("Data inválida ou ausente.")
        if peso_bruto is None or peso_bruto <= 0:
            erros.append("Peso bruto deve ser maior que zero.")
        if peso_liquido is None or peso_liquido <= 0:
            erros.append("Peso líquido calculado está ausente ou inválido.")
        elif peso_bruto is not None and peso_liquido > peso_bruto:
            erros.append("Peso líquido não pode superar o peso bruto.")
        _validar_percentual("Umidade", umidade, erros)
        _validar_percentual("Impureza", impureza, erros)
        _validar_percentual("Defeitos", defeitos, erros)
        normalizados = {
            "propriedade_nome": produtor,
            "cultura": cultura,
            "safra": safra,
            "cadpro_numero": normalizar_cadpro(produtor),
            "classificacao_codigo": "PADRAO",
            "data": data,
            "produtor": produtor,
            "peso_bruto_kg": peso_bruto,
            "peso_liquido_kg": peso_liquido,
            "umidade_percentual": umidade,
            "impureza_percentual": impureza,
            "defeitos_percentual": defeitos,
            "ph": decimal_ou_none(celulas[8].value),
            "destinado_semente": bool(
                normalizar_texto(celulas[9].value)
            ),
        }
        resultados.append(
            _linha_base(
                planilha=sheet.title,
                linha_origem=numero,
                tipo=LinhaImportacao.Tipo.TERCEIROS,
                originais=originais,
                normalizados=normalizados,
                erros=erros,
                advertencias=advertencias,
                associador=associador,
            )
        )
    return resultados


def _validar_container_xlsx(conteudo):
    try:
        with ZipFile(BytesIO(conteudo)) as arquivo_zip:
            entradas = arquivo_zip.infolist()
            if len(entradas) > MAX_ENTRADAS_ZIP:
                raise PlanilhaImportacaoError(
                    "A planilha possui entradas internas em excesso."
                )
            tamanho = sum(item.file_size for item in entradas)
            if tamanho > MAX_CONTEUDO_DESCOMPACTADO_BYTES:
                raise PlanilhaImportacaoError(
                    "O conteúdo descompactado da planilha excede 100 MB."
                )
            if any(item.flag_bits & 0x1 for item in entradas):
                raise PlanilhaImportacaoError(
                    "Planilhas XLSX criptografadas não são suportadas."
                )
    except BadZipFile as exc:
        raise PlanilhaImportacaoError("O arquivo não é um XLSX válido.") from exc


def _ler_arquivo(arquivo):
    nome = Path(getattr(arquivo, "name", "")).name
    if Path(nome).suffix.lower() != ".xlsx":
        raise PlanilhaImportacaoError("Envie um arquivo com extensão .xlsx.")
    conteudo = arquivo.read()
    if not conteudo:
        raise PlanilhaImportacaoError("O arquivo está vazio.")
    if len(conteudo) > MAX_ARQUIVO_BYTES:
        raise PlanilhaImportacaoError("O arquivo excede o limite de 10 MB.")
    _validar_container_xlsx(conteudo)
    return nome, conteudo


def _parsear_workbook(conteudo):
    try:
        workbook = load_workbook(
            BytesIO(conteudo),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        workbook_formulas = load_workbook(
            BytesIO(conteudo),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise PlanilhaImportacaoError(
            "Não foi possível abrir a planilha XLSX."
        ) from exc

    associador = AssociadorGraos()
    linhas = []
    processadas = []
    ignoradas = []
    cabecalhos = {}
    formulas_por_titulo = {
        sheet.title: sheet for sheet in workbook_formulas.worksheets
    }
    try:
        for sheet in workbook.worksheets:
            if sheet.title in PLANILHAS_PRODUCAO:
                cabecalhos[sheet.title] = _validar_cabecalhos(
                    sheet,
                    LinhaImportacao.Tipo.PRODUCAO,
                )
                processadas.append(sheet.title)
                linhas.extend(
                    _parse_producao(
                        sheet,
                        formulas_por_titulo[sheet.title],
                        associador,
                    )
                )
            elif sheet.title == "SAÍDA":
                processadas.append(sheet.title)
                cabecalhos[sheet.title] = _validar_cabecalhos(
                    sheet,
                    LinhaImportacao.Tipo.SAIDA,
                )
                linhas.extend(
                    _parse_saida(
                        sheet,
                        formulas_por_titulo[sheet.title],
                        associador,
                    )
                )
            elif sheet.title == "TERCEIROS":
                processadas.append(sheet.title)
                cabecalhos[sheet.title] = _validar_cabecalhos(
                    sheet,
                    LinhaImportacao.Tipo.TERCEIROS,
                )
                linhas.extend(
                    _parse_terceiros(
                        sheet,
                        formulas_por_titulo[sheet.title],
                        associador,
                    )
                )
            else:
                ignoradas.append(sheet.title)
    finally:
        workbook.close()
        workbook_formulas.close()

    if not linhas:
        raise PlanilhaImportacaoError(
            "Nenhuma linha de produção, saída ou terceiros foi encontrada."
        )
    vistos = {}
    for linha in linhas:
        hash_linha = linha["hash_linha"]
        if hash_linha in vistos:
            linha["advertencias"].append(
                f"Linha potencialmente duplicada de "
                f"{vistos[hash_linha][0]}!{vistos[hash_linha][1]}."
            )
            if linha["status"] == LinhaImportacao.Status.VALIDA:
                linha["status"] = LinhaImportacao.Status.ADVERTENCIA
        else:
            vistos[hash_linha] = (linha["planilha"], linha["linha_origem"])
    return linhas, processadas, ignoradas, cabecalhos


def processar_preview_planilha(*, arquivo, usuario):
    nome, conteudo = _ler_arquivo(arquivo)
    arquivo_sha256 = hashlib.sha256(conteudo).hexdigest()
    existente = LoteImportacao.objects.filter(
        arquivo_sha256=arquivo_sha256
    ).first()
    if existente:
        raise ArquivoImportacaoDuplicadoError(existente)

    linhas, processadas, ignoradas, cabecalhos = _parsear_workbook(conteudo)
    total_erros = sum(
        linha["status"] == LinhaImportacao.Status.ERRO for linha in linhas
    )
    total_advertencias = sum(
        linha["status"] == LinhaImportacao.Status.ADVERTENCIA for linha in linhas
    )
    total_validas = sum(
        linha["status"] == LinhaImportacao.Status.VALIDA for linha in linhas
    )
    total_duplicadas = sum(
        any("potencialmente duplicada" in aviso for aviso in linha["advertencias"])
        for linha in linhas
    )
    status = (
        LoteImportacao.Status.COM_ERROS
        if total_erros
        else LoteImportacao.Status.CONCLUIDO
    )
    try:
        with transaction.atomic():
            lote = LoteImportacao.objects.create(
                arquivo_nome=nome,
                arquivo_tamanho=len(conteudo),
                arquivo_sha256=arquivo_sha256,
                status=status,
                total_planilhas=len(processadas),
                total_linhas=len(linhas),
                total_validas=total_validas,
                total_advertencias=total_advertencias,
                total_erros=total_erros,
                metadados={
                    "planilhas_processadas": processadas,
                    "planilhas_ignoradas": ignoradas,
                    "cabecalhos_reconhecidos": cabecalhos,
                    "total_duplicadas": total_duplicadas,
                    "total_ignoradas": 0,
                    "gera_movimentacoes": False,
                },
                criado_por=usuario,
            )
            objetos = [
                LinhaImportacao(
                    lote_importacao=lote,
                    sequencia=sequencia,
                    **linha,
                )
                for sequencia, linha in enumerate(linhas, start=1)
            ]
            LinhaImportacao.objects.bulk_create(objetos)
    except IntegrityError as exc:
        existente = LoteImportacao.objects.filter(
            arquivo_sha256=arquivo_sha256
        ).first()
        if existente:
            raise ArquivoImportacaoDuplicadoError(existente) from exc
        raise
    return lote
