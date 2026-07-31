from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from apps.graos.models import (
    ArmazemGraos,
    LoteGraos,
    MovimentacaoGraos,
)
from apps.graos.services import saldo_lote
from apps.propriedades.models import Propriedade

from .models import LinhaImportacao, LoteImportacao
from .services import (
    ArquivoImportacaoDuplicadoError,
    PlanilhaImportacaoError,
    processar_preview_planilha,
)


def criar_planilha_teste(nome="preview-soja.xlsx"):
    workbook = Workbook()
    workbook.remove(workbook.active)

    producao = workbook.create_sheet("1")
    producao["C2"] = "Fazenda Modelo"
    producao["C3"] = 10
    producao["F3"] = "SOJA"
    producao["C4"] = "25/26"
    producao.append([])
    producao.append([])
    producao.append([])
    producao["B6"] = "DATA"
    producao["C6"] = "PLACA"
    producao["D6"] = "PESO (Kg)"
    producao["K6"] = "PESO LIQUIDO (Kg)"
    producao["B7"] = date(2026, 2, 1)
    producao["C7"] = "ABC 1D23"
    producao["D7"] = 1000
    producao["E7"] = 14
    producao["F7"] = 1
    producao["G7"] = 0
    producao["H7"] = 80
    producao["K7"] = 950
    producao["B8"] = "data-invalida"
    producao["D8"] = -1
    producao["K8"] = 0

    saida = workbook.create_sheet("SAÍDA")
    saida["C3"] = "SOJA"
    saida["C4"] = "25/26"
    saida["B6"] = "DATA"
    saida["C6"] = "DESTINO"
    saida["G6"] = "CADPRO"
    saida["L6"] = "PESO LIQUIDO (Kg)"
    saida["B7"] = date(2026, 3, 1)
    saida["C7"] = "C.VALE"
    saida["F7"] = "XYZ 9A99"
    saida["G7"] = "Fazenda Modelo"
    saida["H7"] = "Produtor"
    saida["I7"] = "C-100"
    saida["J7"] = "NP-10"
    saida["L7"] = 100

    terceiros = workbook.create_sheet("TERCEIROS")
    terceiros["C3"] = "SOJA"
    terceiros["C4"] = "25/26"
    terceiros["B6"] = "DATA"
    terceiros["C6"] = "PRODUTOR"
    terceiros["E6"] = "PESO (Kg)"
    terceiros["K6"] = "PESO LIQUIDO (Kg)"
    terceiros["B7"] = date(2026, 3, 2)
    terceiros["C7"] = "Terceiro sem cadastro"
    terceiros["E7"] = 500
    terceiros["F7"] = 14
    terceiros["G7"] = 1
    terceiros["H7"] = 0
    terceiros["I7"] = 80
    terceiros["K7"] = 490

    workbook.create_sheet("MENU")
    conteudo = BytesIO()
    workbook.save(conteudo)
    workbook.close()
    return SimpleUploadedFile(
        nome,
        conteudo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


class ImportacaoBase:
    def criar_contexto(self):
        self.usuario = get_user_model().objects.create_user("importador")
        self.propriedade = Propriedade.objects.create(
            nome="Fazenda Modelo",
            municipio="Sorriso",
            uf="MT",
            area_hectares="100",
        )
        self.armazem = ArmazemGraos.objects.create(
            propriedade=self.propriedade,
            nome="Silo Modelo",
            capacidade_kg="100000",
        )
        self.lote_graos = LoteGraos.objects.create(
            armazem=self.armazem,
            codigo="SOJA-25-26",
            cultura="Soja",
            safra="2025/2026",
        )


class ImportacaoServiceTests(ImportacaoBase, TestCase):
    def setUp(self):
        self.criar_contexto()

    def test_preview_persiste_lote_linhas_e_associacoes_sem_movimentar(self):
        saldo_antes = saldo_lote(self.lote_graos)
        lote = processar_preview_planilha(
            arquivo=criar_planilha_teste(),
            usuario=self.usuario,
        )
        self.assertEqual(lote.status, LoteImportacao.Status.COM_ERROS)
        self.assertEqual(lote.total_planilhas, 3)
        self.assertEqual(lote.total_linhas, 4)
        self.assertEqual(lote.total_validas, 2)
        self.assertEqual(lote.total_advertencias, 1)
        self.assertEqual(lote.total_erros, 1)
        self.assertFalse(lote.metadados["gera_movimentacoes"])
        self.assertEqual(lote.metadados["total_duplicadas"], 0)
        self.assertIn("1", lote.metadados["cabecalhos_reconhecidos"])

        producao = lote.linhas.get(planilha="1", linha_origem=7)
        self.assertEqual(producao.status, LinhaImportacao.Status.VALIDA)
        self.assertEqual(producao.propriedade, self.propriedade)
        self.assertEqual(producao.lote_graos, self.lote_graos)
        self.assertEqual(
            producao.associacao,
            LinhaImportacao.Associacao.LOTE_GRAOS,
        )
        self.assertEqual(
            producao.dados_normalizados["peso_liquido_kg"],
            "950",
        )
        self.assertEqual(
            producao.dados_normalizados["classificacao_codigo"],
            "PADRAO",
        )

        invalida = lote.linhas.get(planilha="1", linha_origem=8)
        self.assertEqual(invalida.status, LinhaImportacao.Status.ERRO)
        self.assertGreaterEqual(len(invalida.erros), 3)

        terceiros = lote.linhas.get(planilha="TERCEIROS")
        self.assertEqual(
            terceiros.status,
            LinhaImportacao.Status.ADVERTENCIA,
        )
        self.assertTrue(terceiros.advertencias)
        self.assertEqual(MovimentacaoGraos.objects.count(), 0)
        self.assertEqual(saldo_lote(self.lote_graos), saldo_antes)

    def test_linha_repetida_e_sinalizada_sem_ser_descartada(self):
        arquivo = criar_planilha_teste()
        conteudo = BytesIO(arquivo.read())
        workbook = load_workbook(conteudo)
        producao = workbook["1"]
        for coluna in range(2, 12):
            producao.cell(9, coluna).value = producao.cell(7, coluna).value
        repetida = BytesIO()
        workbook.save(repetida)
        workbook.close()

        lote = processar_preview_planilha(
            arquivo=SimpleUploadedFile(
                "preview-com-duplicidade.xlsx",
                repetida.getvalue(),
            ),
            usuario=self.usuario,
        )
        linha = lote.linhas.get(planilha="1", linha_origem=9)
        self.assertEqual(linha.status, LinhaImportacao.Status.ADVERTENCIA)
        self.assertTrue(
            any("potencialmente duplicada" in aviso for aviso in linha.advertencias)
        )
        self.assertEqual(lote.total_linhas, 5)
        self.assertEqual(MovimentacaoGraos.objects.count(), 0)

    def test_hash_impede_reimportacao_do_mesmo_arquivo(self):
        primeiro = criar_planilha_teste()
        conteudo = primeiro.read()
        lote = processar_preview_planilha(
            arquivo=SimpleUploadedFile("preview-soja.xlsx", conteudo),
            usuario=self.usuario,
        )
        with self.assertRaises(ArquivoImportacaoDuplicadoError) as contexto:
            processar_preview_planilha(
                arquivo=SimpleUploadedFile("preview-soja.xlsx", conteudo),
                usuario=self.usuario,
            )
        self.assertEqual(contexto.exception.lote, lote)
        self.assertEqual(LoteImportacao.objects.count(), 1)

    def test_rejeita_extensao_invalida(self):
        arquivo = SimpleUploadedFile(
            "dados.csv",
            b"conteudo",
            content_type="text/csv",
        )
        with self.assertRaisesMessage(PlanilhaImportacaoError, ".xlsx"):
            processar_preview_planilha(
                arquivo=arquivo,
                usuario=self.usuario,
            )

    def test_rejeita_xlsx_corrompido(self):
        arquivo = SimpleUploadedFile(
            "dados.xlsx",
            b"nao-e-um-zip",
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        with self.assertRaisesMessage(PlanilhaImportacaoError, "XLSX válido"):
            processar_preview_planilha(
                arquivo=arquivo,
                usuario=self.usuario,
            )


    def test_rejeita_layout_sem_cabecalhos_obrigatorios(self):
        arquivo = criar_planilha_teste()
        conteudo = BytesIO(arquivo.read())
        workbook = load_workbook(conteudo)
        workbook["1"]["K6"] = None
        invalida = BytesIO()
        workbook.save(invalida)
        workbook.close()

        with self.assertRaisesMessage(
            PlanilhaImportacaoError,
            "cabecalhos obrigatorios",
        ):
            processar_preview_planilha(
                arquivo=SimpleUploadedFile(
                    "layout-invalido.xlsx",
                    invalida.getvalue(),
                ),
                usuario=self.usuario,
            )


class ImportacaoApiTests(ImportacaoBase, APITestCase):
    def setUp(self):
        self.criar_contexto()
        self.client.force_authenticate(self.usuario)

    def test_autenticacao_e_obrigatoria(self):
        self.client.force_authenticate(None)
        self.assertEqual(
            self.client.get("/api/importacoes/lotes/").status_code,
            401,
        )
        self.assertEqual(
            self.client.get("/api/importacoes/linhas/").status_code,
            401,
        )
        self.assertEqual(
            self.client.post("/api/importacoes/lotes/preview/", {}).status_code,
            401,
        )

    def test_preview_e_consulta_auditavel(self):
        resposta = self.client.post(
            "/api/importacoes/lotes/preview/",
            {"arquivo": criar_planilha_teste()},
            format="multipart",
        )
        self.assertEqual(resposta.status_code, 201, resposta.data)
        self.assertIn("no-store", resposta["Cache-Control"])
        self.assertIn("private", resposta["Cache-Control"])
        lote_id = resposta.data["lote"]["id"]
        self.assertEqual(resposta.data["lote"]["total_linhas"], 4)
        self.assertEqual(len(resposta.data["linhas_preview"]), 4)
        self.assertFalse(resposta.data["preview_limitado"])

        linhas = self.client.get(
            f"/api/importacoes/linhas/?lote={lote_id}&tipo=producao"
        )
        self.assertEqual(linhas.status_code, 200)
        self.assertEqual(len(linhas.data), 2)
        self.assertEqual(MovimentacaoGraos.objects.count(), 0)

    def test_upload_duplicado_retorna_409_e_lote_existente(self):
        arquivo = criar_planilha_teste()
        conteudo = arquivo.read()
        primeira = self.client.post(
            "/api/importacoes/lotes/preview/",
            {"arquivo": SimpleUploadedFile("preview-soja.xlsx", conteudo)},
            format="multipart",
        )
        repetida = self.client.post(
            "/api/importacoes/lotes/preview/",
            {"arquivo": SimpleUploadedFile("preview-soja.xlsx", conteudo)},
            format="multipart",
        )
        self.assertEqual(primeira.status_code, 201)
        self.assertEqual(repetida.status_code, 409)
        self.assertEqual(
            repetida.data["lote_existente"],
            primeira.data["lote"]["id"],
        )

    def test_lotes_e_linhas_sao_imutaveis_pela_api(self):
        lote = processar_preview_planilha(
            arquivo=criar_planilha_teste(),
            usuario=self.usuario,
        )
        linha = lote.linhas.first()
        self.assertEqual(
            self.client.patch(
                f"/api/importacoes/lotes/{lote.id}/",
                {"status": "concluido"},
                format="json",
            ).status_code,
            405,
        )
        self.assertEqual(
            self.client.delete(
                f"/api/importacoes/linhas/{linha.id}/"
            ).status_code,
            405,
        )

    def test_openapi_documenta_endpoint_de_preview(self):
        resposta = self.client.get("/api/schema.json")
        self.assertEqual(resposta.status_code, 200)
        self.assertIn(
            "/importacoes/lotes/preview/",
            resposta.data["paths"],
        )
